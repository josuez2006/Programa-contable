from openpyxl import Workbook
import os

def data_to_excel_convert(data: list, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active

    debit_counter = 1
    credit_counter = 1

    for entry in data:

        if entry.is_negative == True:
            letter_position = 65
            number_position = debit_counter

            data: list[list]  = entry.how_to_print_in_excel(letter_position, number_position)

            for info in data:
                ws[chr(info[1]) + str(info[2])] = info[0]

            debit_counter += entry.get_number_of_breaks_in_excel()

        else:
            letter_position = 75
            number_position = credit_counter

            data: list[list]  = entry.how_to_print_in_excel(letter_position, number_position)

            for info in data:
                ws[chr(info[1]) + str(info[2])] = info[0]

            credit_counter += entry.get_number_of_breaks_in_excel()



    file_path = f'{output_path}/Resumen_Contable.xlsx'
    wb.save(file_path)
    os.system(file_path)